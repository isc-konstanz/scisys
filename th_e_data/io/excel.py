# -*- coding: utf-8 -*-
"""
    th-e-sim.io.excel
    ~~~~~~~~~~~~~~~~~


"""
import os
import logging
import warnings
import pandas as pd

from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Font, Side
from tables import NaturalNameWarning
from copy import copy

warnings.filterwarnings('ignore', category=NaturalNameWarning)
logger = logging.getLogger(__name__)


def write_excel(summary, data_frames, dir: str = 'data', file: str = 'summary.xlsx'):
    border_side = Side(border_style=None)
    border = Border(top=border_side,
                    right=border_side,
                    bottom=border_side,
                    left=border_side)

    if not os.path.isabs(file):
        summary_path = os.path.join(dir, file)
    else:
        summary_path = file
    if not summary_path.endswith('.xlsx'):
        summary_path += '.xlsx'

    summary_dir = os.path.dirname(summary_path)
    if not os.path.isdir(summary_dir):
        os.makedirs(summary_dir, exist_ok=True)

    with pd.ExcelWriter(summary_path, engine='openpyxl') as summary_writer:
        summary.to_excel(summary_writer, sheet_name='Summary', float_format="%.2f")  # , encoding='utf-8-sig')
        summary_book = summary_writer.book

        for data_key, data in data_frames.items():
            data.to_excel(summary_writer, sheet_name=data_key)  # , encoding='utf-8-sig')
            data_sheet = summary_book[data_key]
            for column in range(1, len(data_sheet[1])):
                data_column_width = data.iloc[:, column - 1].apply(lambda s: len(str(s))).max()
                data_column_width = max(data_column_width, len(str(data_sheet[1][column].value)))
                data_sheet.column_dimensions[get_column_letter(column + 1)].width = data_column_width + 2

        # Set column width and header coloring
        for data_sheet in summary_book:
            if data_sheet.title == 'Summary':
                data_sheet.delete_rows(3, 1)
                header_len = summary.columns.nlevels
            else:
                header_len = data_frames[data_sheet.title].columns.nlevels

            header_font = Font(name="Calibri Light", size=12, color='333333')

            for column in range(len(data_sheet[header_len])):
                for header_row in range(1, header_len + 1):
                    header_cell = data_sheet[header_row][column]
                    if '\n' in str(header_cell.value):
                        header_alignment = copy(header_cell.alignment)
                        header_alignment.wrapText = True
                        # header_alignment.vertical = 'center'
                        # header_alignment.horizontal = 'center'
                        header_cell.alignment = header_alignment
                        data_sheet.row_dimensions[header_row].height = 33

                    header_cell.font = header_font
                    header_cell.border = border

                if data_sheet.title == 'Summary' or column == 0:
                    data_column_width = 0
                    for data_cell in data_sheet[get_column_letter(column + 1)]:
                        data_cell.border = border
                        data_column_width = max(data_column_width, len(str(data_cell.value)))
                    data_sheet.column_dimensions[get_column_letter(column + 1)].width = data_column_width + 2
